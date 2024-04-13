#!/usr/bin/env python3
# -*- coding: utf-8 -*-


# ana dosya bu bunu çalıştır 
"""
Created on Thu Jan 27 01:14:03 2022

@author: beyza sayraci
lobe segmentation train code
"""

from simple_multi_unet_model import multi_unet_model #Uses softmax 
import os
import numpy as np
from matplotlib import pyplot as plt
import tensorflow as tf
import random
#from tensorflow.random import set_seed
from keras import backend as K
from generate_data import imageLoader
import constants
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import pandas as pd
import datetime
import glob
from shutil import move
from  model import *
#from tensorflow.python.keras.utils.multi_gpu_utils import multi_gpu_model
import keras
from keras.callbacks import ModelCheckpoint,TensorBoard, EarlyStopping
import matplotlib
from matplotlib import pyplot as plt
import glob
import random
#from tensorflow.random import set_seed
#import tensorflow_addons as tfa
# from tensorflow.keras.utils import multi_gpu_model
#from tensorflow.python.keras.utils.multi_gpu_utils import multi_gpu_model
from shutil import move
#import segmentation_models as sm
from sklearn.preprocessing import MinMaxScaler
from tensorflow.keras.optimizers import SGD, RMSprop, Adam, Adadelta, Adagrad, Adamax, Nadam, Ftrl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from keras import backend as K
import constants
from keras.models import load_model
matplotlib.use('Agg') # use to save the figure without showing the figure 

DATASET        = constants.DATASET                    # get dataset path
train_img_dir  = constants.train_img_dir              # get train image path
train_mask_dir = constants.train_mask_dir             # get train mask path
results_dir    = constants.results_dir
tmp_dir        = constants.tmp_dir                    # get temp path
models_dir     = constants.models_dir                 # get model path
    

blind_test_img_dir  = constants.blind_test_img_dir     # get blind test image path 
blind_test_mask_dir = constants.blind_test_mask_dir    # get blind test mask path  
blind_test_file     = constants.blind_test_file        # get blind test file path  

val_img_dir  = constants.val_img_dir    # get validation image path
val_mask_dir = constants.val_mask_dir   # get validation mask path

epochs              = constants.EPOCHS                 # get epoch 
batch_size          = constants.BATCH_SIZE             # get batch size
ARCH                = constants.ARCH                   # get architecture
OPTIMIZER           = constants.OPTIMIZER              # get optimizer
learning_rate       = constants.LEARNING_RATE          # get learning rat
MOMENTUM            = constants.MOMENTUM
KERNEL_INITIALIZER  = constants.KERNEL_INITIALIZER     # get kernel initializer
SHOW_PLT            = constants.SHOW_PLT               # get SHOW_PLT variable
# IS_GPU              = constants.IS_GPU                 # get IS_GPU variable


SEED_NUMBER = constants.SEED_NUMBER  # get seed number
os.environ['PYTHONHASHSEED'] = str(SEED_NUMBER)
np.random.seed(SEED_NUMBER)
#random.seed(SEED_NUMBER)
tf.random.set_seed(SEED_NUMBER)


# tf.compat.v1.debugging.set_log_device_placement(True)         
os.environ["CUDA_VISIBLE_DEVICES"] = "0"
print("GPU: {}, CPU: {}".format(tf.config.list_physical_devices('GPU'),tf.config.list_physical_devices('CPU')))
tf.config.run_functions_eagerly(True)

def dice_coef(y_true, y_pred, smooth=1):
      

        y_true_f     = K.flatten(y_true)                                                   # get vector of the y_true matrix
        y_pred_f     = K.flatten(y_pred)                                                   # get vector of the y_pred matrix
        intersection = K.sum(y_true_f * y_pred_f)                                          # get intersection of the two vectors
        return (2. * intersection + smooth) / (K.sum(y_true_f) + K.sum(y_pred_f) + smooth) # return dice coefficient 

def iou_score(y_true, y_pred):
       
        def f(y_true, y_pred):
            intersection = (y_true * y_pred).sum()                     # get intersection y_true and y_pred
            union        = y_true.sum() + y_pred.sum() - intersection  # get union y_true and y_pred
            x            = (intersection + 1e-15) / (union + 1e-15)    # get iou score
            x            = x.astype(np.float32)                        # convert iou score from float64 to float32
            return x                                                   # return iou score
        return tf.numpy_function(f, [y_true, y_pred], tf.float32)      # return iou score
    

def my_resize(image,shape_=(256,256)):
    img_h = image.shape[0]
    img_w = image.shape[1]
    rimg_h = shape_[0]
    rimg_w = shape_[1]

    delta_h = int(img_h/rimg_h)
    delta_w = int(img_w/rimg_w)

    image = np.delete(image,range(0,img_h,delta_h),1)
    image = np.delete(image,range(0,img_w,delta_w),0)
    
    return image       


SIZE_X = 256
SIZE_Y = 256
n_classes=3




train_mask_list = sorted(os.listdir(train_mask_dir))   
train_img_list = sorted(os.listdir(train_img_dir))
validation_mask_list = sorted(os.listdir(val_mask_dir))
validation_image_list = sorted(os.listdir(val_img_dir))

train_img_datagen = imageLoader(train_img_dir, train_img_list, train_mask_dir, train_mask_list, batch_size)
val_img_datagen = imageLoader(val_img_dir, validation_image_list, val_mask_dir, validation_mask_list, batch_size)



def get_optimizer():
        """
        @ This method returns the optimizers.
        @ params : 
             @@ 
        @ return : 
            @@
        """
        global optimizer
        
        if(OPTIMIZER == 'Adam'):                                                           # check optimizers
            optimizer = Adam(learning_rate=learning_rate)                             # get optimizer
        elif(OPTIMIZER == 'SGD'):                                                          # check optimizers
            optimizer = SGD(learning_rate=learning_rate,momentum=MOMENTUM)       # get optimizer
        elif(OPTIMIZER == 'RMSprop'):                                                      # check optimizers
            optimizer = RMSprop(learning_rate=learning_rate,momentum=MOMENTUM)   # get optimizer
        elif(OPTIMIZER == 'Adadelta'):                                                     # check optimizers
            optimizer = Adadelta(learning_rate=learning_rate)                         # get optimizer
        elif(OPTIMIZER == 'Adagrad'):                                                      # check optimizers
            optimizer = Adagrad(learning_rate=learning_rate)                          # get optimizer
        elif(OPTIMIZER == 'Adamax'):                                                       # check optimizers
            optimizer = Adamax(learning_rate=learning_rate)                           # get optimizer
        elif(OPTIMIZER == 'Nadam'):                                                        # check optimizers
            optimizer = Nadam(learning_rate=learning_rate)                            # get optimizer
        elif(OPTIMIZER == 'Ftrl'):                                                         # check optimizers
            optimizer = Ftrl(learning_rate=learning_rate)                             # get optimizer
        else:
            raise AssertionError("Please select from the list of the optimizer. {} is not valid optimizer".format(OPTIMIZER)) # raise the error

def get_model():
    
        global model
        
        if ARCH.__eq__("attention_network"):                                                                                                 # check architecture for resunet
            model  =attention_network(n_class=3,IMG_HEIGHT=256, IMG_WIDTH=256, IMG_CHANNELS=1)                                                              # generate resune
            model = unet_2d_model_128_128(n_class=3,IMG_HEIGHT=256, IMG_WIDTH=256, IMG_CHANNELS=1,kernel_initializer=KERNEL_INITIALIZER)  # generate unet model
        elif ARCH.__eq__("128"):                                                                                                       # check architecture for unet
            model = unet_2d_model_128_128(n_class=3,IMG_HEIGHT=256, IMG_WIDTH=256, IMG_CHANNELS=1,kernel_initializer=KERNEL_INITIALIZER)
        elif ARCH.__eq__("256"):                                                                                                       # check architecture for unet
            model = unet_2d_model_256_256(n_class=3,IMG_HEIGHT=256, IMG_WIDTH=256, IMG_CHANNELS=1,kernel_initializer=KERNEL_INITIALIZER)  # generate unet model
        elif ARCH.__eq__("DeeplabV3Plus_vgg19"):
           model = DeeplabV3Plus_vgg19(n_class=3,IMG_HEIGHT=256, IMG_WIDTH=256, IMG_CHANNELS=1)          # generate unet model
        elif ARCH.__eq__("DeeplabV3Plus_vgg16"):
           model = DeeplabV3Plus_vgg16(n_class=3,IMG_HEIGHT=256, IMG_WIDTH=256, IMG_CHANNELS=1)          # generate unet model
        elif ARCH.__eq__("DeepLabV3Plus_resnet50"):
            model = DeeplabV3Plus_resnet50(n_class=3,IMG_HEIGHT=256, IMG_WIDTH=256, IMG_CHANNELS=1)          # generate unet model
        elif ARCH.__eq__("dense_unet_2d"):
            model = dense_unet_2d(n_class=3,IMG_HEIGHT=256, IMG_WIDTH=256, IMG_CHANNELS=1)          # generate unet model
        elif ARCH.__eq__("MultiUnet"):                                                                                                 # check architecture for resunet
            #resunet = ResUnet(IMG_HEIGHT=256, IMG_WIDTH=256, IMG_CHANNELS=1)                                                              # generate resunet model
            #model = resunet.get_model(3)  
            model =   multi_unet_model(n_classes=3, IMG_HEIGHT=256, IMG_WIDTH=256, IMG_CHANNELS=1)                                                                               # get the model
            # model = load_model("DATASETS/Lobe_Segmen/DeepLabV3Plus_resnet50/results/logs_300_epoch_64_batch_multi_unet_modelBwBetul usdllytfhn_arch_84.2_iou_score/models/saved_model-259-259-0.76-0.86.hdf5",custom_objects={"dice_coef":dice_coef, "iou_score":iou_score})
        elif ARCH.__eq__("UNet_vgg16"):                                                                                                   # check architecture for vgg16-unet
            model = UNet_vgg16(IMG_HEIGHT=256, IMG_WIDTH=256, IMG_CHANNELS=1,kernel_initializer=KERNEL_INITIALIZER)                  # generate vgg16-unet model                         
        elif ARCH.__eq__("UNet_vgg19"):                                                                                                   # check architecture for vgg16-unet
            model = UNet_vgg19(IMG_HEIGHT=256, IMG_WIDTH=256, IMG_CHANNELS=1,kernel_initializer=KERNEL_INITIALIZER)                  # generate vgg16-unet model     
        elif ARCH.__eq__("TernausNet16"):                                                                                                   # check architecture for vgg16-unet
            model = ternausNet16(n_class=3,IMG_HEIGHT=256, IMG_WIDTH=256, IMG_CHANNELS=1)                  # generate vgg16-unet model                         
        else: 
            raise AssertionError("Please select from the list of the arhitecture. {} is not valid arhitecture".format(ARCH))         # raise the error"
   
def model_compile_and_fit():
       
        global model, optimizer, history
        model.compile(optimizer=optimizer, loss='categorical_crossentropy', metrics=['accuracy', dice_coef, iou_score]) # metrik olarak IOU score, DSC score(dice diye de geçebilir), mIOU score, precision, Sensitivity and specificity ...

        print(model.summary())                                                                               # print the summary
        print(model.input_shape)                                                                             # print the input shape
        print(model.output_shape)                                                                            # print the output shape

        check_point_dir = DATASET + results_dir +  tmp_dir +  "saved_model-{epoch:02d}-{epoch:02d}-{iou_score:.2f}-{val_iou_score:.2f}.hdf5"    # get check point path
        checkpoint      = ModelCheckpoint(check_point_dir, monitor='iou_score', verbose=1, save_best_only=False, mode='max')   # generate model check point for save the model at the each epoch
        callbacks       = [checkpoint,
                        # EarlyStopping(patience=10, monitor='val_loss',restore_best_weights=False),
                        TensorBoard(log_dir=DATASET + results_dir +  tmp_dir + 'logs'),
                        # ReduceLROnPlateau(patience=10)
                        ] # get callbacks
        steps_per_epoch     = len(train_img_list)//batch_size 
        val_steps_per_epoch = len(validation_image_list)//batch_size  

        history    = model.fit(train_img_datagen,
                          steps_per_epoch=steps_per_epoch,
                          epochs=epochs,
                          verbose=1,
                          validation_data=val_img_datagen,
                          validation_steps=val_steps_per_epoch,
                          callbacks=[callbacks],
                          shuffle=False)       
        history     = history.history       
   
def train(i_train=None,i_test=None,k_fold=False):

     os.environ['PYTHONHASHSEED'] = str(SEED_NUMBER)
     np.random.seed(SEED_NUMBER)
     tf.random.set_seed(SEED_NUMBER)         
     

     get_optimizer()   
     get_model()       
    
     model_compile_and_fit()                                                            
     save_model()                                                                                         
    

def append_to_excel(fpath, df):
        """
        @ This method appends the history to the xlsx file
        @ params : 
            @@ fpath : the file path :: string
            @@ df    : the dataframe :: pandas.Dataframe
        @ returns : 
            @@ 
        """
       
        book  = load_workbook(fpath)                               # load xlsx file
        sheet = book.get_sheet_by_name(book.get_sheet_names()[0])  # get the sheet
        
        for row in dataframe_to_rows(df, index=False, header=False):    # load the dataframe into the sheet
            sheet.append(row)                                           # append the row to the sheet
        
        book.save(fpath)   # save the modified excel at the desired locatio
        
def get_dict_index(dictionary,value):
      list_ = [i for i, d in enumerate(dictionary) if d==value ] # find index in the dictionary
      return list_[-1]  
  
def make_path():
    
        global path
        
        path = path + "/"                   
        while True:                              
            if not os.path.exists(path):       
                 break                           
            path = path                     
            path = path + "_hash_"+str(hash(random.randint(0, 100000000000)))+str(hash(random.randint(0, 100000000000)))+"/" # add random hash to the path
            
        os.mkdir(path)    # generate path
        path = path  # set the local path to the global path

def save_model():
    global history, path
    
    max_iou_score        = max(history['iou_score'])                                          
    max_val_iou_score    = max(history['val_iou_score'])                                      
    max_dsc_score        = max(history['dice_coef'])                                         
    max_val_dsc_score    = max(history['val_dice_coef'])
    
    max_iou_score_in =  get_dict_index(history['iou_score'],max_iou_score) + 1 
    max_val_iou_score_in = get_dict_index(history['val_iou_score'],max_val_iou_score) + 1
    max_iou_score        = round(max_iou_score*100,2)                                            
    max_val_iou_score    = round(max_val_iou_score*100,2)                                          
    max_dsc_score        = round(max_dsc_score*100,2)                                         
    max_val_dsc_score    = round(max_val_dsc_score*100,2)      

    sub_path = "logs_{}_epoch_{}_batch_{}_arch_{}_iou_score".format(epochs,batch_size,ARCH,max_iou_score) 
    path     = DATASET+"results/" + sub_path                                                                  
    make_path()                               
   
    file_name = path + sub_path +".xlsx" 
    
    df  = pd.DataFrame(history)              
    df1 = pd.DataFrame({"Epoch":[epochs],
                        "Batch":[batch_size],
                        "Architecture":[ARCH],
                        "Optimizer":[OPTIMIZER],
                        "Learning_Rate":[learning_rate],
                        "Momentum" :[MOMENTUM],
                        "Kernel_Initializer" :[KERNEL_INITIALIZER],
                        "Seed_Number":[SEED_NUMBER],                        
                        "Max_IOU_Score":[max_iou_score],                     
                        "Max_Val_IOU_Score":[max_val_iou_score],
                        "Max_DSC":[max_dsc_score],
                        "Max_Val_DSC":[max_val_dsc_score],
                        "Dataset":[DATASET.replace("/", "")],
                        "Date":[datetime.datetime.now().strftime("%d.%m.%Y - %H.%M")],
                        "Annotions":[constants.ANNOTIONS]}) 

    df = df.append(df1)                                                         
    df.to_excel(file_name)                                                      
    append_to_excel(DATASET+results_dir+"all_results.xlsx", df1) 
     
    os.mkdir(path+models_dir)                                                       
    saved_models = sorted(glob.glob(DATASET + results_dir + tmp_dir +"*.hdf5")) 
    logs_path = DATASET + results_dir + tmp_dir + "logs" 
    move(logs_path, path + models_dir)
        
    for _,path_ in enumerate(saved_models):                                                     
        _,file_name = os.path.split(path_)                                                                            
        in_ = int(file_name.split('-')[1])                                                      
        if(max_iou_score_in != in_ and max_val_iou_score_in != in_):                               
            os.remove(path_)                                                                     
        else:
            move(path_, path + models_dir)   
            
def save_plot(fig_name_loss,fig_name_iou,is_train=True):
    
        global history
    
        if (not hasattr("history")):                                                                # check the history variable if exists
            raise AssertionError("Please call first model_compile_and_fit() or blind_test_each_epoch()") # raise the error

        loss      = history['loss']       # get training loss
        
        iou_score = history['iou_score']  # get training iou score
        
        fig_ = [[loss,'','Testing loss','','Testing Loss',fig_name_loss],[iou_score,'','Testing IOU Score','','Testing IOU Score',fig_name_iou]]  # get labels and scores

        if(is_train):                                          # check the training status                   
            val_loss      = history['val_loss']      # get validation loss
            val_iou_score = history['val_iou_score'] # get validation iou score
        
            fig_ =  [[loss,val_loss,'Training loss','Validation loss','Training and Validation Loss',fig_name_loss],[iou_score,val_iou_score,'Training IOU Score','Validation IOU Score','Training and Validation IOU Score',fig_name_iou]]   # get labels and scores
    
                                                                                                             
        for index, [line1,line2, label1,label2, title,fig_name] in enumerate(fig_) :   # enumarate fig_ list
            epochs = range(1, len(line1) + 1)                                          # generate epoch range

            plt.figure()                                    # generate new figure for plotting
            plt.plot(epochs, line1, 'y', label=label1)      # plot training loss or iou score
            if(is_train):                                   # check the training status
                plt.plot(epochs, line2, 'r', label=label2)  # plot validation loss  or iou score
            plt.title(title)                                # set the title to the plot
            plt.xlabel('Epochs')                            # set the label to the plot
            if(index == 0):                                 # check the index
                plt.ylabel('Loss')                          # set the label to the plot
            else:
                plt.ylabel('IOU Score')                     # set the label to the plot
            plt.legend()                                    # set the legend to the plot
            plt.savefig(fig_name)                           # save the figure 
            if SHOW_PLT:                               # check to show figure
                plt.draw()                                  # show figure

                              
            
# def show_predicted_image(blind_test_img,blind_test_mask,pred_blind_test):
    
#       global fig
#       fig.canvas.flush_events()            # clear the figure
#       sp1 = fig.add_subplot(231)           # generate subplot
#       sp1.set_title('Testing Image')            # set the title to the plot
#       sp1.imshow(blind_test_img, cmap='gray')   # show the testing original images
#       sp2 = fig.add_subplot(232)           # generate subplot
#       sp2.set_title('Testing Label')            # set the title to the plot
#       sp2.imshow(blind_test_mask)               # show the testing mask
#       sp3 = fig.add_subplot(233)           # generate subplot
#       sp3.set_title('Prediction on test image') # set the title to the plot
#       sp3.imshow(pred_blind_test)               # show the testing predicted image
#       plt.draw()                                # show the figure
#       plt.pause(0.01)                           # pause 10 ms
# def initialize_blind_test():
#         global blind_test_img_batch,blind_test_mask_batch
  
#         blind_test_img_list    = os.listdir(blind_test_img_dir)   
#         blind_test_mask_list   = os.listdir(blind_test_mask_dir)
     
#         blind_test_img_datagen = imageLoader(blind_test_img_dir, blind_test_img_list, blind_test_mask_dir, blind_test_mask_list, len(blind_test_img_list)) 
       
#         blind_test_img_batch, blind_test_mask_batch = blind_test_img_datagen.__next__()    
            
# def blind_test_each_epoch():
    
#         global history, model, saved_models
    
#         initialize_blind_test()     # initialize blind test

       
#         saved_models = sorted(glob.glob(DATASET + results_dir + blind_test_file + "/" +  models_dir + "*.hdf5")) # get saved models
#         all_history = list()                                                                                                         # generate the list to save the all history
        
#         for _,file in enumerate(saved_models): 
#             print("load model",_)           
#             model   = load_model(file,compile=True,custom_objects="categorical_crossentropy")                                       
#             loss      = history['loss']      # get testing loss
#             iou_score = history['iou_score'] # get testing iou score
            
#             all_history.append([iou_score,loss]) # add the history to the all_history list
            
#         all_history  = np.array(all_history)                                        # convert list to np.array
#         history =  {"loss":all_history[:][0],"iou_score":all_history[:][1]}    # get history
#         # save_plot(DATASET + results_dir+blind_test_file+"/"+ blind_test_file +"_testing_loss.png",DATASET + results_dir+blind_test_file+"/" + blind_test_file + "_testing_iou.png",is_train=False) # save plots

# def find_best_model():
#     global saved_models
#     raise NotImplementedError("This method is not implemented!")
#     saved_models = sorted(glob.glob(DATASET + results_dir + blind_test_file + "/" +  models_dir + "*.hdf5")) # get saved models

    
# def blind_test_last_epoch():
    
#         global model,blind_test_img_batch,blind_test_mask_batch,fig,show_predicted_image
#         initialize_blind_test()  # initialize blind test    
#         model   = load_model(DATASET+results_dir+blind_test_file+'/'+blind_test_file+'.hdf5',compile=True,custom_objects="categorical_crossentropy")  # load the model that is at each eposh. For predictions you do not need to compile the model, so ...
#         history = model.evaluate(blind_test_img_batch,blind_test_mask_batch, verbose=1,return_dict=True)                                          # evaluate the metrics for last epoch
#         print(history) 
        
#         preds_blind_test = model.predict(blind_test_img_batch,verbose=1)
#         preds_blind_test = (preds_blind_test > 0.5).astype(np.uint8)         
#         preds_blind_test = np.squeeze(preds_blind_test,axis=3)                
#         quit = True                                                           
#         plt.ion()                                                            
#         fig = plt.figure(figsize=(12, 8))                               
#         plt.draw()                                                           
#         while quit: # infinity loop
     
#             index = random.randint(0, len(preds_blind_test)-1) 
                    
#             blind_test_img  = blind_test_img_batch[index]   
#             blind_test_mask = blind_test_mask_batch[index] 
#             pred_blind_test = preds_blind_test[index]            
          
#             show_predicted_image(blind_test_img,blind_test_mask,pred_blind_test)  
                    
#             quit = bool(input("Would you please press the enter to quit or any other character for continue?"))  # get the character from the user to quit

train()


# from keras.models import load_model
# # import segmentation_models_3D as sm
# from generate_data import imageLoader
# import numpy as np

# my_model = load_model('saved_model-299-299-0.98-0.98.hdf5', 
#                       compile=False)

# from keras.metrics import MeanIoU

# batch_size=8 #Check IoU for a batch of images
# test_img_datagen = imageLoader(val_img_dir, validation_image_list, 
#                                 val_mask_dir, validation_mask_list, batch_size)

# #Verify generator.... In python 3 next() is renamed as __next__()
# test_image_batch, test_mask_batch = test_img_datagen.__next__()

# test_mask_batch_argmax = np.argmax(test_mask_batch, axis=4)
# test_pred_batch = my_model.predict(test_image_batch)
# test_pred_batch_argmax = np.argmax(test_pred_batch, axis=4)

# n_classes = 6
# IOU_keras = MeanIoU(num_classes=n_classes)  
# IOU_keras.update_state(test_pred_batch_argmax, test_mask_batch_argmax)
# print("Mean IoU =", IOU_keras.result().numpy())

# #############################################
# #Predict on a few test images, one at a time
# #Try images: 
# img_num = 82

# test_img = np.load("/home/eva/Desktop/Beyza/Lobe_Segmen/DATASETS/Lobe_Segmen/DeepLabV3Plus_resnet50/input_data_splitted/val/images"+str(img_num)+".npy")

# test_mask = np.load("/home/eva/Desktop/Beyza/Lobe_Segmen/DATASETS/Lobe_Segmen/DeepLabV3Plus_resnet50/input_data_splitted/val/masks"+str(img_num)+".npy")
# test_mask_argmax=np.argmax(test_mask, axis=3)

# test_img_input = np.expand_dims(test_img, axis=0)
# test_prediction = my_model.predict(test_img_input)
# test_prediction_argmax=np.argmax(test_prediction, axis=4)[0,:,:,:]


# print(test_prediction_argmax.shape)
# print(test_mask_argmax.shape)
# print(np.unique(test_prediction_argmax))


#Plot individual slices from test predictions for verification
# from matplotlib import pyplot as plt
# import random

# #n_slice=random.randint(0, test_prediction_argmax.shape[2])
# n_slice = 55
# plt.figure(figsize=(12, 8))
# plt.subplot(231)
# plt.title('Testing Image')
# plt.imshow(test_img[:,:,n_slice,1], cmap='gray')
# plt.subplot(232)
# plt.title('Testing Label')
# plt.imshow(test_mask_argmax[:,:,n_slice])
# plt.subplot(233)
# plt.title('Prediction on test image')
# plt.imshow(test_prediction_argmax[:,:, n_slice])
# plt.show()




# if __name__=="__main__":
    # train()                                    
# iou_score = history['iou_score']
# val_iou_score = history['val_iou_score']
# epochs = range(1, len(iou_score) + 1)
# plt.plot(epochs, iou_score, 'y', label='Training iou score')
# plt.plot(epochs, val_iou_score, 'r', label='Validation iou score')
# plt.title('Training and validation iou score')
# plt.xlabel('Epochs')
# plt.ylabel('Loss')
# plt.legend()
# plt.show()


# dice_coef = history['dice_coef']
# val_dice_coef = history['val_dice_coef']
# epochs = range(1, len(dice_coef) + 1)
# plt.plot(epochs, dice_coef, 'y', label='Training dice coef')
# plt.plot(epochs, val_dice_coef, 'r', label='Validation dice coef')
# plt.title('Training and validation dice coef')
# plt.xlabel('Epochs')
# plt.ylabel('Loss')
# plt.legend()
# plt.show()

# print("max iou_score is = ", (max(iou_score) * 100.0), "% at",iou_score.index(max(iou_score))+1)
# print("max val_iou_score is = ", (max(val_iou_score) * 100.0), "% at",val_iou_score.index(max(val_iou_score))+1)

# print("max dice_coef is = ", (max(dice_coef) * 100.0), "% at",dice_coef.index(max(dice_coef))+1)
# print("max val_dice_coef is = ", (max(val_dice_coef) * 100.0), "% at",val_dice_coef.index(max(val_dice_coef))+1)

# loss = history['loss']
# val_loss = history['val_loss']
# epochs = range(1, len(loss) + 1)
# plt.plot(epochs, loss, 'y', label='Training loss')
# plt.plot(epochs, val_loss, 'r', label='Validation loss')
# plt.title('Training and validation loss')
# plt.xlabel('Epochs')
# plt.ylabel('Loss')
# plt.legend()
# plt.show()

# accuracy = history['accuracy']
# val_accuracy = history['val_accuracy']

# plt.plot(epochs, accuracy, 'y', label='Training Accuracy')
# plt.plot(epochs, val_accuracy, 'r', label='Validation Accuracy')
# plt.title('Training and validation Accuracy')
# plt.xlabel('Epochs')
# plt.ylabel('Accuracy')
# plt.legend()
# plt.show()
